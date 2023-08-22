import pandas as pd
import numpy as np
import openpyxl
import os
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.metrics import mean_absolute_error
import json
import datetime
import csv
import time
import math

from typing import Union
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

#global variables
day = '10'
month = '10'
year = '2022'

app = FastAPI()
origins = [
    "http://localhost:3000",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class Event(BaseModel):
    start: int
    end: int
    label: Union[str, None] = None

class Options(BaseModel):
    category: str
    length: int
    perDay: int
    recommendations: int
    selectedDays: list[int]
    color: str
    
class TrainingData(BaseModel):
    previousweek1: list[Event] = []
    previousweek2: list[Event] = []
    previousweek3: list[Event] = []
    previousweek4: list[Event] = []

class RequestCurrentWeek(BaseModel):
    currentWeek: list[Event] = []
    trainingData: TrainingData
    options: Options
    selectedWeek: str

class RequestAllEvents(BaseModel):
    trainingData: TrainingData
    selectedWeek: str

def dayToInt(day):
    match day:
        case "Monday":
            return 0
        case "Tuesday":
            return 1
        case "Wednesday":
            return 2
        case "Thursday":
            return 3
        case "Friday":
            return 4
        case "Saturday":
            return 5
        case "Sunday":
            return 6 


def get_n_time_value(n, d, var):
    i = 0
    for j in d:
        if i == n:
            return j[var]
        i = i +1
        
def format_request(rawdata):
    # print(rawdata)
    data = []
    for i in rawdata:
        i.start = int(str(i.start)[:10])
        i.end = int(str(i.end)[:10])
        datestart_time = datetime.datetime.fromtimestamp(i.start)
        dateend_time = datetime.datetime.fromtimestamp(i.end)
        data.append({'start': datestart_time.hour*60 + datestart_time.minute, 'end': dateend_time.hour*60 + dateend_time.minute, 'day': dayToInt(datestart_time.strftime("%A"))})
    #print('Data:', data)
    return data

def convert_csv(dict_item, fields):
    with open('test4.csv', 'w', newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames = fields)
        writer.writeheader()
        writer.writerows(dict_item)
    return 'test4.csv'

def read_csv(dir):
    with open(dir, 'r') as f:
        csv_reader = csv.reader(f)
        for line in csv_reader:
            print(line)

def create_at_table(events):
    a_t_table = []
    timeArray = np.arange(0,1440,30)
    eventIndex = 0
    timeIndex = 0
    dayIndex = 0
    # print('a_t_events:', events)
    # Checking if an empty array gets sent
    if events:
        isEventsChecked = False
    else:
        isEventsChecked = True

    while(timeIndex < 48 and dayIndex < 7):
        #print(eventIndex, timeIndex, dayIndex)
        # If current time = start Time of event and the day = event.day and all events have not been checked
        if not isEventsChecked and timeArray[timeIndex] == events[eventIndex]['start'] and dayIndex == events[eventIndex]['day']:
            # Check to make sure the end does not go into a new day
            #print('Skipping from: ', events[eventIndex]['start'], ' to ', events[eventIndex]['end']) 
            while (eventIndex+1) < len(events) and events[eventIndex]['end'] == events[eventIndex+1]['start']:
                if (events[eventIndex]['start'] >events[eventIndex]['end']):
                    dayIndex += 1
                eventIndex += 1
            
            if events[eventIndex]['end'] > events[eventIndex]['start']:
                # Change current time to end of event if it does not go into new day
                timeIndex = int(events[eventIndex]['end']/30)
            else:
                # Change current time and day if it goes into a new day
                #print("new day!")
                timeIndex = int(events[eventIndex]['end']/30)
                dayIndex += 1
            # Check for next event start time if next event exists
            if (eventIndex+1) != len(events):
                eventIndex += 1
                # If it doesn't then we dont need to worry about these if statements anymore
            else:
                isEventsChecked = True
        # Add this 30 min time slot into a_t_table
        a_t_table.append({'time': timeArray[timeIndex], 'day': dayIndex})
        # Check to see if we have reached the end of the day to reset time and increase day index
        if timeArray[timeIndex] == 1410:
            #print('new day!')
            dayIndex += 1
            timeIndex = 0
            continue
        timeIndex += 1
    # print("Final: ", a_t_table)
    return a_t_table

# def possible_timeslots(y, dataset):
#     index = 0
#     count = 0
#     index2 = 0
#     possible_times = []
#     #print(len(dataset))
#     for i in dataset:
#         data = []
#         if index+y  < len(dataset):
#             for x in range(y+1):    
#                 data.append(dataset[index])
#                 index = index + 1
#             count = count + 1
#             index = count
#         #print(data)
#         increment = 30
#         for l in range(y):
#             if int(get_n_time_value(l+1, data, 'time') or 0) != int(get_n_time_value(l, data, 'time') or 0) + increment or int(get_n_time_value(l+1, data, 'day') or 0) != int(get_n_time_value(l, data, 'day') or 0):
#                 #print("not valid")
#                 break
#             if l == y-1:
#                 possible_times.append(data)
#                 #print("YES VALID: " + str(data))
#                 break
#     # print("possible times: " + str(possible_times))
#     # print(len(possible_times))
#     return possible_times

def possible_timeslots(y, dataset):
    # print('possible_timeslots(y): ', y)
    # print('possible_timeslots(dataset): ', dataset)
    index = 0
    del dataset[0]
    possible_times = []
    index=0
    data_length = len(dataset)-1
    # print('dataset:', dataset)
    # print(len(dataset))
    while(index < data_length):
        # print('new loop!', 'index: ', index, ' data_lenght: ', data_length)
        data = []
        for x in range(y+1):
            current_index = index + x
            next_index = current_index + 1
            if next_index >= data_length:
                index = next_index
                break
            # print('current_index: ', dataset[current_index][0], 'next_index: ', (dataset[next_index][0]))
            if int(dataset[current_index][0])+30 == int(dataset[next_index][0]):
                data.append({'time': dataset[current_index][0], 'day': dataset[current_index][1], 'probability': dataset[current_index][2]})
            elif dataset[current_index][0] == 1410 and dataset[next_index][0] == 0:
                data.append({'time': dataset[current_index][0], 'day': dataset[current_index][1], 'probability': dataset[current_index][2]})
            elif x != y:
                # print('x: ', x, 'y :', y)
                index = current_index
        # print(len(data), ' ', y)
        if len(data) == y + 1:
            possible_times.append(data)
        index += 1
    # print('possible_times:', possible_times)
    return possible_times

# def get_probability_timeslots(probability_dir, possible_timeslots):
#     prob_total = ['Probability']
#     index = 1
#     with open(probability_dir) as f:
#         probability_list = list(csv.reader(f))
#     # print('possible_timeslots: ', possible_timeslots)
#     # print('probability_list:', probability_list)
#     for i in possible_timeslots:
#         # print('i:', i)
#         prob = []
#         for j in i:
#             # print('probability_lsit[index][2]: ', probability_list[index][2])
#             prob.append(probability_list[index][2])
#         index = index +1
#         prob_total.append(total_probability(prob))
#     return prob_total

# THIS IS HOW I DID THE CODE BEFORE BECAUSE OF THE STUPID WAY YOU NEED TO DECLARE A LIST OF LIST (I WILL EXPLAIN IN CLASS SMH)
# def get_probability_timeslots(possible_timeslots, top_n):
#     prob_total = [] # Holds total probabilities of each possible_timeslots[index]
#     events_total_prob = [] # Same data structure as events total, however only holding probilities to easily get the top probabilities 
#     events_total = [] # Holds all the events, sorted into 7 days with data structure [[day_total],[day_total],[day_total],[day_total],[day_total],[day_total],[day_total]]
#     print(events_total)
#     day_total = [] # Holds all events for a particular day
#     prev_day = 0
#     index = 0
#     # Iterate through the possible timeslots and add their probabilities together
#     # print(possible_timeslots)
#     for i in possible_timeslots:
#         prob = []
#         start_day = 0
#         end_day = 0
#         start = 0
#         end = 0
#         # print('i:', i)
#         for j in range(len(i)):
#             if j == 0: #if its the first in the iteration
#                 start = i[j]['time']
#                 start_day = i[j]['day']
#             elif j == len(i)-1: #if its the last in the iteration
#                 end = i[j]['time']
#                 end_day = i[j]['day']
#             # print('j[probability]', i[j]['probability'])
#             prob.append(i[j]['probability'])
#         # print('start_day:', int(start_day))
#         # For some reason I cant appened directly to the array so I create day_total which I append to events_total instead of creating a matrix for events_total
#         # I then append it to the events_total to create the matrix I want but I really don't like that I had to do it this way
#         # When reading array.append documentation looks like can do something like events_total[index].append (I can but it appends for all arrays instead of the index array being referenced)
#         if int(start_day) != int(prev_day): # If it is a new day
#             # print('start_day: ', start_day, ' prev_day: ', prev_day, ' len(possible_timeslots): ', len(possible_timeslots), ' index + 1', index+1)
#             events_total.append(day_total)
#             events_total_prob.append(prob_total)
#             day_total = []
#             prob_total = []
#         elif len(possible_timeslots) == index+1: # Else if it is the last event
#             day_total.append({'start': start, 'end': end, 'start_day': int(start_day), 'end_day': int(end_day), 'probability': total_probability(prob)})
#             prob_total.append(total_probability(prob)) 
#             events_total.append(day_total)
#             events_total_prob.append(prob_total)
#         # Otherwise keep appending events to day_total
#         day_total.append({'start': start, 'end': end, 'start_day': int(start_day), 'end_day': int(end_day), 'probability': total_probability(prob)})
#         prob_total.append(total_probability(prob))
#         prev_day = start_day
#         index += 1
#     # return 'Test'
#     return get_top_events(events_total_prob, events_total, top_n)

# def get_top_events(events_total_prob, events_total, top_n):
#     top_events = []
#     top_events_index = []
#     index = 0
#     for i in events_total_prob:
#         top_events_index.append(top_x(i, top_n))
#     # print('top_events_index: ', top_events_index)
#     for i in top_events_index:
#         for j in i:
#             # print('events_total[i][j]: ', events_total[index][j])
#             top_events.append(events_total[index][j])
#         index += 1
#     print('top_events: ', top_events)
#     return top_events

def get_probability_timeslots_v2(possible_timeslots, top_n, label_name, color):
    prob_total = [[] for i in range(7)]
    events_total = [[] for i in range(7)]
    print('events_total(empty array):', events_total)
    index = 1
    for i in possible_timeslots:
        prob = []
        start_day = 0
        end_day = 0
        start = 0
        end = 0
        for j in range(len(i)):
            if j == 0:
                start = i[j]['time']
                start_day = i[j]['day']
            elif j == len(i)-1:
                end = i[j]['time']
                end_day = i[j]['day']
            # print('j[probability]', i[j]['probability'])
            prob.append(i[j]['probability'])
        index = index +1
        prob_total[int(start_day)].append(total_probability(prob))
        events_total[int(start_day)].append({'start': start, 'end': end, 'start_day': int(start_day), 'end_day': int(end_day), 'probability': total_probability(prob)})
        # print('len(prob_total)', len(prob_total))
        # print('prob_total:', prob_total)
    # print('events_total[0]: ', events_total[0]) 
    return get_top_events_v2(prob_total, events_total, top_n, label_name, color)

def get_top_events_v2(prob_total, events_total, top_n, label_name, color):
    top_events = [[] for i in range(7)]
    for index, i in enumerate(prob_total):
        prob_total[index] = top_x(i, top_n)
    print('prob_total: ', prob_total)
    for index, i in enumerate(prob_total):
        for j in i:
            print('events_total[index][j]: ', events_total[index][j])
            start = events_total[index][j]['start']
            end = events_total[index][j]['end']
            start_day = events_total[index][j]['start_day']
            end_day = events_total[index][j]['end_day']
            label_name = label_name
            event_color = color[:-1] + ", 0.1)" # Evan: transparant color rgba(..., 0.1)
            top_events[index].append({
                'start': create_epoch(start, start_day),
                'end': create_epoch(end, end_day),
                'start_day': start_day,
                'end_day': end_day,
                'name': label_name,
                'category': label_name,
                'timed': True,
                'recommend': True,
                'color': event_color,
                'probability': events_total[index][j]['probability']
            })
    print('top_events: ', top_events)
    return json.dumps(top_events)
     


def total_probability(array):
    total = 0
    count = 0
    for i in array:
        count = count + 1
        total = total + float(i)
    return total/count

def top_x(array, top_n):
    print(array)
    top_x_idx = np.argsort(array)[-top_n:]
    top_x_index = [i for i in top_x_idx]
    # top_x_values = [array[i] for i in top_x_idx]
    # del top_x_index[-1]
    # print('top_x_values', top_x_values)
    # print('top_x_index:', top_x_index)
    return top_x_index

class KNN:  
    
    def __init__(self, name, dataset_dir, dataset_pred):
        self.name = name
        self.data = pd.read_excel(dataset_dir)
        # Just confirming, since you pass csv into predict_probability you dont need to initialise they hey?
        self.pred = pd.read_csv(dataset_pred)
        
    def accuracy_test(self): #returns the accuracy of the KNN algorithm on a particular dataset
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        dataset_train, dataset_test, category_train, category_test = train_test_split(dataset, category, test_size=0.20, random_state=13)
        scaler = StandardScaler() #fit the data into the algorithm 
        scaler.fit(dataset_train)
        dataset_train = scaler.transform(dataset_train)
        dataset_test = scaler.transform(dataset_test)
        classifier = KNeighborsClassifier(n_neighbors=3) #Create the model
        classifier.fit(dataset_train, category_train) #Train the model
        return classifier.score(dataset_test, category_test) # return accruacy of a model based on a 20:80 dataset split
    
    def train_model(self, num_neighbours):
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        model = KNeighborsClassifier(num_neighbours)
        model.fit(dataset, category)
        return model
    
    def predict(self, model, data_pred):
        df = pd.read_excel(data_pred)
        dataset_pred=df.drop(['Category'], axis = 1)
        predictions = model.predict(dataset_pred)
        print(predictions)
        return predictions

    def predict_probaility(self, model, dataset_pred, label_name):
        df = pd.read_csv(dataset_pred)
        #print(model.classes_)
        target_column = 0
        for i in model.classes_:
            if i == label_name:
                break
            target_column = target_column+1
        proba = [label_name]
        for i in model.predict_proba(df):
            index = 0
            for j in i:
                if index == target_column:
                    proba.append(j)
                index = index+1
        with open(dataset_pred, 'r') as fi:
            lines = [[i.strip() for i in line.strip().split(',')] \
                    for line in fi.readlines()]
            print("LINESSSSS: ", lines)
        new_lines = [line + [str(proba[i])] for i, line in enumerate(lines)]
        prediction_table_dir = 'prediction_table.csv'
        with open(prediction_table_dir, 'w') as fo:
            for line in new_lines:
                fo.write(','.join(line) + '\n')
        return prediction_table_dir

    def predict_single_probaility(self, model, timeslot):
        return model.predict_proba(timeslot)
    
    def find_optimal_n(self): #Need to test this function to see if it runs slow as it retrains a model 40 times
        error = []
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        dataset_train, dataset_test, category_train, category_test = train_test_split(dataset, category, test_size=0.20, random_state=13)
        scaler = StandardScaler() #fit the data into the algorithm 
        scaler.fit(dataset_train)
        dataset_train = scaler.transform(dataset_train)
        dataset_test = scaler.transform(dataset_test)
        for i in range(1, 40):
            knn = KNeighborsClassifier(n_neighbors=i)
            knn.fit(dataset_train, category_train)
            pred_i = knn.predict(dataset_test)
            mae = knn.score(dataset_test, category_test)
            error.append(mae)
        return [error.index(np.max(error)), np.amax(error)]

    def find_optimal_n2(self): #Need to test this function to see if it runs slow as it retrains a model 40 times
        error = []
        dataset=self.data.drop(['Category'], axis = 1) #Seperate the prediction column
        category=self.data['Category']
        dataset_train, dataset_test, category_train, category_test = train_test_split(dataset, category, test_size=0.20, random_state=13)
        scaler = StandardScaler() #fit the data into the algorithm 
        scaler.fit(dataset_train)
        dataset_train = scaler.transform(dataset_train)
        dataset_test = scaler.transform(dataset_test)
        for i in range(1, 40):
            knn = KNeighborsClassifier(n_neighbors=i)
            knn.fit(dataset_train, category_train)
            pred_i = knn.predict(dataset_test)
            mae = knn.score(dataset_test, category_test)
            error.append(mae)
        return error
# C:\Users\ndeo\Documents\Projects\Uni\SIS\knnfastapi
def test():
    dataset = '.\TheoreticalData.xlsx' 
    prediction_dataset = 'C:\\Users\\Calvin\\Documents\\SES_STUDIO\\KNN\\TestMon.xlsx'
    calvinKNN = KNN("Calvin", dataset, prediction_dataset, "Sleep")
    print(calvinKNN.accuracy_test())
    model = calvinKNN.train_model(5)
    #calvinKNN.predict(model)
    #calvinKNN.predict_probaility(model, prediction_dataset)
    print(calvinKNN.find_optimal_n())

def get_all_events_week(): #need to fix this function to read csv instead of excel
    dataset = 'TheoreticalData.xlsx' 
    pd = 'prediction_table.csv' 
    prediction_dataset = 'TestMon.xlsx'
    calvinKNN = KNN("Calvin", dataset, pd)
    model = calvinKNN.train_model(5)
    wb = openpyxl.load_workbook('TestMon.xlsx')
    sheet1 = wb['Sheet1']
    count = 2
    for i in calvinKNN.predict(model, prediction_dataset):
        sheet1.cell(count, column=3).value = i
        count = count + 1
    wb.save('TestMon.xlsx')
    events = []
    event_type_old = ""
    event = {"start": 0, "end":0, "start_day":0, "end_day":0,  "name": ""}
    for row in sheet1.iter_rows(min_row=2):
        if event_type_old == row[2].value and event['end'] == row[0].value:
            event['end'] = int(event['end']) +30
            event['end_day'] = row[1].value
        else:
            events.append(event)
            event = {"start":row[0].value, 
                    "end": int(row[0].value)+30, 
                    "start_day":row[1].value,
                    "end_day":row[1].value,
                    "name":row[2].value}
        event_type_old = row[2].value
    events.append(event)
    events.pop(0)
    events[:] = [d for d in events if d.get('name') != "Blank"]
    # convert to epoch
    epoch_events =[]
    for i in events:
        #print(i['name'])
        epoch_events.append({"start": create_epoch(i['start'], i['start_day']), "end": create_epoch(i['end'], i['end_day']),  "name": i['name'], "timed": True, 'probability': 1})
    print(epoch_events)
    return epoch_events
        

    #print(events)
        # print(event_type_old)
        # print(event)
        # if event_type_old == j2.value:
        #     event = {"start":j[0].value, 
        #             "end":j[1].value+30, 
        #             "name":j[2].value}

def assign_globals(x):
    date = x.split('.')
    global day 
    day = date[0]
    global month 
    month = date[1]
    global year 
    year = date[2]


@app.get("/")
def read_root():
    return {"Hello": "World"}


@app.get("/items/{item_id}")
def read_item(item_id: int, q: Union[str, None] = None):
    item_id = item_id+ 1
    return {"item_id": item_id, "q": q}

@app.get("/knn")
def run_test():
    return test2()

@app.get("/stats/classification_report") # Classifcation report (how accurate it is at predicting each variable ) spliting the dataset 80-20
async def run_test():
    df = pd.read_excel('.\TheoreticalData.xlsx')
    x=df.drop(['Category'], axis = 1)
    y=df['Category']
    X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.20, random_state=13)
    classifier = KNeighborsClassifier(n_neighbors=1)
    classifier.fit(X_train, y_train)
    y_pred = classifier.predict(X_test)
    report = classification_report(y_test, y_pred, output_dict=True)
    df = pd.DataFrame(report).transpose()
    print(df)
    #print(classification_report(y_test, y_pred))
    #print(confusion_matrix(y_test, y_pred))
    return df

@app.get("/stats/confusion_matrix") # Classifcation report (how accurate it is at predicting each variable ) spliting the dataset 80-20
async def run_test():
    df = pd.read_excel('.\TheoreticalData.xlsx')
    x=df.drop(['Category'], axis = 1)
    y=df['Category']
    X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.20, random_state=13)
    classifier = KNeighborsClassifier(n_neighbors=1)
    classifier.fit(X_train, y_train)
    y_pred = classifier.predict(X_test)
    report = confusion_matrix(y_test, y_pred)
    #print(classification_report(y_test, y_pred))
    data = []
    for i in report:
        data.append(str(i))
    print(data)
    return data

@app.get("/stats/dataframe_stats") # time and day variables -> column count, mean of each category, std, min, max , 25%, 50%, 75%
async def run_test():
    df = pd.read_excel('.\TheoreticalData.xlsx')
    x=df.drop(['Category'], axis = 1)
    y=df['Category']
    return df.describe().T

@app.get("/stats/neighbour_accuracy") # Array of neighbour number to accuracy -> index refers to number of neighbours e.g array[0] = 1 nieghbour = 93% accuracy
async def create_item():
    dataset = '.\TheoreticalData.xlsx' 
    pd = 'prediction_table.csv' 
    calvinKNN = KNN("Calvin", dataset, pd)
    return calvinKNN.find_optimal_n2()

@app.post("/calendar/all")
async def create_item(request: RequestAllEvents):
    #print(request.selectedWeek)
    assign_globals(request.selectedWeek)
    return get_all_events_week()

@app.post("/calendar")
async def create_item(request: RequestCurrentWeek):
    assign_globals(request.selectedWeek)
    # Uncomment the following line and replace training_data variables for hard coded schedule
    # dataset = '.\TheoreticalData.xlsx' 
    training_data = create_training_data(request.trainingData.previousweek1, request.trainingData.previousweek2, request.trainingData.previousweek3, request.trainingData.previousweek4)
    convert_prediction = create_at_table(format_request(request.currentWeek))
    prediction_dataset = convert_csv(convert_prediction, ['time', 'day'])
    # Do you need to initialise prediction_dataset?
    calvinKNN = KNN("Calvin", training_data, prediction_dataset) # Change to training_data for dynamic training data
    model = calvinKNN.train_model(5)
    prob_dir = calvinKNN.predict_probaility(model, prediction_dataset, request.options.category)
    with open(prob_dir) as f:
        probability_list = list(csv.reader(f))
    #print('probability_list:', probability_list)
    p_t = possible_timeslots(request.options.length, probability_list)
    print('possible_timeslots: ', p_t)
    return get_probability_timeslots_v2(p_t, request.options.recommendations, request.options.category, request.options.color)
    return append_final_events_v2(top_values,"Sleep")

def create_training_data(one, two, three, four):
    wb = openpyxl.load_workbook(r'c:trainingdata.xlsx')
    ws = wb['Sheet1']
    for row in ws:
        for cell in row:
            cell.value = None
    ws.cell(1, column=1).value = "time"
    ws.cell(1, column=2).value = "day"
    ws.cell(1, column=3).value = "Category"
    count = 2
    rawdata = []
    rawdata.append(one)
    rawdata.append(two)
    rawdata.append(three)
    rawdata.append(four)
    for j in rawdata:
        data = []
        for i in j:
            i.start = int(str(i.start)[:10])
            i.end = int(str(i.end)[:10])
            datestart_time = datetime.datetime.fromtimestamp(i.start)
            dateend_time = datetime.datetime.fromtimestamp(i.end)
            data.append({'start': datestart_time.hour*60 + datestart_time.minute, 'end': dateend_time.hour*60 + dateend_time.minute, 'day': dayToInt(datestart_time.strftime("%A")), 'label': i.label})
        for i in data:
            start = i['start']
            while(start < i['end']):
                ws.cell(count, column=1).value = start
                ws.cell(count, column=2).value = i['day']
                ws.cell(count, column=3).value = i['label']
                count = count + 1
                start = start + 30
        blanks = create_at_table(format_request(j))
        for i in blanks:
            ws.cell(count, column=1).value = i['time']
            ws.cell(count, column=2).value = i['day']
            ws.cell(count, column=3).value = "Blank"
            count = count+1
    wb.save('trainingdata.xlsx')
    return "trainingdata.xlsx"


# def append_final_events(top_x, possible_timeslots, label_name):
#     final_events = []
#     # startweek = "18.09.2022 00:00:00"
#     # create_epoch(startweek)
#     for i in top_x:
#         index = 0
#         start = 0
#         start_day = 0
#         end = 0
#         end_day = 0
#         for j in possible_timeslots[i]:
#             if index == 0:
#                 start = j['time']
#                 start_day=j['day']
#             if index == len(possible_timeslots[i]) -1:
#                 end = (j['time'])
#                 end_day=j['day']
#             index = index + 1
#         final_events.append({'start': create_epoch(start, start_day), 'end': (create_epoch(end, end_day)+1800000), 'name':label_name,'timed': True})
#     Events = json.dumps(final_events)
#     return Events
    # print(possible_timeslots[25])

# def append_final_events_v2(top_x, label_name):
#     # print('top_x:', top_x)
#     final_events_total = []
#     final_events_day = []
#     prev_day = 0
#     # startweek = "18.09.2022 00:00:00"
#     # create_epoch(startweek)
#     # I don't like how I did this as well. Might come back and implement a better way (just came up with it at 3:33am haha) however this one works for now 
#     for i in top_x:
#         if prev_day != i['start_day']:
#             final_events_total.append(final_events_day)
#             final_events_day = []
#         elif i == top_x[len(top_x)-1]:
#             final_events_day.append({'start': create_epoch(i['start'], i['start_day']), 'end': (create_epoch(i['end'], i['end_day'])+1800000), 'name':label_name,'timed': True})
#             final_events_total.append(final_events_day)
#             break
#         final_events_day.append({'start': create_epoch(i['start'], i['start_day']), 'end': (create_epoch(i['end'], i['end_day'])+1800000), 'name':label_name,'timed': True})
#         prev_day = i['start_day']
#     print('final_events_total: ', final_events_total)
#     Events = json.dumps(final_events_total)
#     return Events
#     # print(possible_timeslots[25])

def create_epoch(t, d):
    # Dont forget to change the 26 in day to a variable
    hour = math.floor(int(t)/60)
    minute = int(t)%60
    second = 0
    if hour == 24: # fixes my code where i sent the end time to 1430 which is 0 for next day)
        hour = 0
        d = d + 1
    total_time = str(day)+ "." + str(month) + "." + str(year) + " " + str(hour) +":" + str(minute) + ":"+ str(second)
    #print(total_time)
    pattern = '%d.%m.%Y %H:%M:%S'
    epoch = int(time.mktime(time.strptime(total_time, pattern)))*1000 + 86400000*d
    return epoch




# [Event(start=1663524000000, end=1663567200000, label='Event #4'), Event(start=1663554600000, end=1663599600000, label='asdfasdfasdf'),
#  Event(start=1663783200000, end=1663795800000, label='Save Event'), Event(start=1663695000000, end=1663714800000, label='Test Realtime'), 
#  Event(start=1663515000000, end=1663520400000, label='Test Realtime 2')]

# Function - Finding the highest probabilty events for x number of days in the respective week
# Parameters in - rawdata, x (num of days), y (num of hours per day), dataset table
# Parameters out - Highest_Probabilty_Per_Day
# Ensure time is sorted beforehand to make checking consective times easier
# @app.get("/findHighestProbabilty")
# def findHighestProbabilty(rawdata, dataset, x, y):
    # 1. dataset to train the model
    # 2. inverse the rawdata to get the a_t_table(as it sends timeslots that are full to get)
    # {
    #    put the a_t_table into the model to get the prediction_table
    # }
    # 3. Find possible timeslots that correlate to y into an array object[{day:"Monday","60", "90"}, (day:"Tuesday" "90","120")] index = day
    # { 
    #    Iterate over prediction_table 
    #    if current_day != previous day
    #       count= 0
    #       max_probability = -1
    #       if count*30 >= y 
    #           find probability sum of consecutive days
    #           if current_probability > max_probability  
    #               max_probability = current_probability 
    #               Highest_Probabilty_Per_Day[current_day] = max_probability
    # }
    # 4.Find probability of each array group
    # remove min value of Highest_Probabilty_Per_Day[7] until index = x
    # Highest_Probabilty_Per_Day[7] = {"highest probability of Monday",  "highest probability of Tuesday"}
    # return Highest_Probabilty_Per_Day

#To Run Local Python File
#.\venv\Scripts\activate
#install module dependencies
#py main.py

#To Build Docker File
#.\venv\Scripts\activate
#docker build -t python-fastapi .
#docker run -p 8000:8000 python-fastapi
#docker run -d --publish 8888:5000 python-fastapi


if __name__ == '__main__':
    #get_all_events_week()
    uvicorn.run(app, port=int(os.environ.get("PORT", 5000)), host="0.0.0.0")